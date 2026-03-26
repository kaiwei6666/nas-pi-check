import os
import re
import csv
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

NAS_PATH = r"\\192.168.1.132\startup\107年創基所有資料-1101021版-OK\1-深耕管考平台資料-OK\1.計畫成果(含工作會議或活動)-OK"

SKIP_DIRS = {
    "#recycle",
    "@eaDir",
    "$RECYCLE.BIN",
    "System Volume Information"
}

SUPPORTED_EXTENSIONS = {".txt", ".csv", ".log", ".docx", ".xlsx", ".pdf"}

# 用來存放偵測到個資的檔案資訊
pii_files = []

def detect_pii_types(text):
    pii_types = []

    patterns = {
        "身分證字號": r'[A-Z][12]\d{8}',
        "手機號碼": r'09\d{8}',
        "Email": r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+',
    }

    for pii_name, pattern in patterns.items():
        if re.search(pattern, text):
            pii_types.append(pii_name)

    return pii_types

def read_text_file(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def read_docx_file(file_path):
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text]
    return "\n".join(paragraphs)

def read_xlsx_file(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    all_text = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                all_text.append(" ".join(row_text))

    return "\n".join(all_text)

def read_pdf_file(file_path):
    reader = PdfReader(file_path)
    all_text = []

    for page in reader.pages:
        text = page.extract_text()
        if text:
            all_text.append(text)

    return "\n".join(all_text)

def extract_text(file_path):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        return None

    if ext in {".txt", ".csv", ".log"}:
        return read_text_file(file_path)
    elif ext == ".docx":
        return read_docx_file(file_path)
    elif ext == ".xlsx":
        return read_xlsx_file(file_path)
    elif ext == ".pdf":
        return read_pdf_file(file_path)

    return None

def scan_file(file_path):
    try:
        text = extract_text(file_path)
        if not text:
            return []

        pii_types = detect_pii_types(text)
        return pii_types

    except Exception as e:
        print(f"讀檔失敗: {file_path} | {e}")
        return []

def handle_walk_error(err):
    print(f"略過無法存取路徑: {err}")

def export_csv():
    if not pii_files:
        print("📄 沒有偵測到任何含個資檔案，不輸出 CSV。")
        return

    output_file = "pii_report.csv"

    with open(output_file, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["檔名", "副檔名", "完整路徑", "大小(KB)", "偵測到的個資類型"])

        for item in pii_files:
            writer.writerow([
                item["name"],
                item["ext"],
                item["path"],
                item["size_kb"],
                "、".join(item["pii_types"])
            ])

    print(f"\n📄 已輸出報表：{output_file}")

def scan_nas():
    if not os.path.exists(NAS_PATH):
        print("❌ NAS 路徑不存在")
        return

    print("🔍 開始掃描 NAS...\n")

    for root, dirs, files in os.walk(NAS_PATH, topdown=True, onerror=handle_walk_error):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]

        for file_name in files:
            file_path = os.path.join(root, file_name)
            _, ext = os.path.splitext(file_path)

            if ext.lower() not in SUPPORTED_EXTENSIONS:
                continue

            print(f"掃描中: {file_path}")

            pii_types = scan_file(file_path)

            if pii_types:
                print(f"⚠️ 發現個資: {file_path}")
                print(f"   類型: {', '.join(pii_types)}\n")

                try:
                    size_kb = round(os.path.getsize(file_path) / 1024, 2)
                except Exception:
                    size_kb = "N/A"

                pii_files.append({
                    "name": os.path.basename(file_path),
                    "ext": ext.lower(),
                    "path": file_path,
                    "size_kb": size_kb,
                    "pii_types": pii_types
                })

    print("\n========== 掃描結果 ==========")
    if pii_files:
        for i, item in enumerate(pii_files, start=1):
            print(f"{i}. {item['path']} | 類型: {', '.join(item['pii_types'])}")
        print(f"\n總共找到 {len(pii_files)} 個含個資檔案")
    else:
        print("沒有偵測到含個資的檔案")

    export_csv()

if __name__ == "__main__":
    scan_nas()