import os
import re
import csv
from typing import Callable, List, Dict, Any
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

NAS_PATH = r"\\192.168.1.132\startup"

SKIP_DIRS = {
    "#recycle",
    "@eaDir",
    "$RECYCLE.BIN",
    "System Volume Information",
    "_restricted",
}

SUPPORTED_EXTENSIONS = {".txt", ".csv", ".log", ".docx", ".xlsx", ".pdf"}


def detect_pii_types(text: str) -> List[str]:
    pii_types = []
    patterns = {
        "身分證字號": r'[A-Z][12]\d{8}',
        "手機號碼": r'09\d{8}',
        "Email": r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+',
    }

    for pii_name, pattern in patterns.items():
        if re.search(pattern, text):
            pii_types.append(pii_name)

    return pii_types


def read_text_file(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def read_docx_file(file_path: str) -> str:
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text]
    return "\n".join(paragraphs)


def read_xlsx_file(file_path: str) -> str:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    all_text = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                all_text.append(" ".join(row_text))

    return "\n".join(all_text)


def read_pdf_file(file_path: str) -> str:
    reader = PdfReader(file_path)
    all_text = []

    for page in reader.pages:
        text = page.extract_text()
        if text:
            all_text.append(text)

    return "\n".join(all_text)


def extract_text(file_path: str) -> str | None:
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        return None

    if ext in {".txt", ".csv", ".log"}:
        return read_text_file(file_path)
    if ext == ".docx":
        return read_docx_file(file_path)
    if ext == ".xlsx":
        return read_xlsx_file(file_path)
    if ext == ".pdf":
        return read_pdf_file(file_path)

    return None


def scan_file(file_path: str) -> List[str]:
    try:
        text = extract_text(file_path)
        if not text:
            return []
        return detect_pii_types(text)
    except Exception as e:
        print(f"讀檔失敗: {file_path} | {e}")
        return []


def handle_walk_error(err: OSError, progress_callback: Callable[[str], None] | None = None) -> None:
    msg = f"略過無法存取路徑: {err}"
    if progress_callback:
        progress_callback(msg)
    else:
        print(msg)



def run_scan(config: Dict[str, Any], progress_callback: Callable[[str], None]) -> List[Dict[str, Any]]:
    scan_root = config.get("nas_smb_root") or NAS_PATH

    if not os.path.exists(scan_root):
        raise FileNotFoundError(f"掃描路徑不存在：{scan_root}")

    results: List[Dict[str, Any]] = []
    progress_callback("🔍 開始掃描...")

    def _walk_error(err: OSError) -> None:
        handle_walk_error(err, progress_callback)

    for root, dirs, files in os.walk(scan_root, topdown=True, onerror=_walk_error):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]

        for file_name in files:
            file_path = os.path.join(root, file_name)
            _, ext = os.path.splitext(file_path)

            if ext.lower() not in SUPPORTED_EXTENSIONS:
                continue

            progress_callback(f"掃描中: {file_path}")
            pii_types = scan_file(file_path)

            if pii_types:
                try:
                    size_kb = round(os.path.getsize(file_path) / 1024, 2)
                except Exception:
                    size_kb = "N/A"

                results.append({
                    "name": os.path.basename(file_path),
                    "ext": ext.lower(),
                    "path": file_path,
                    "size_kb": size_kb,
                    "pii_types": pii_types,
                    "action_result": "偵測到個資"
                })

    return results


def export_csv(results: List[Dict[str, Any]], output_file: str = "pii_report.csv") -> None:
    if not results:
        print("📄 沒有偵測到任何含個資檔案，不輸出 CSV。")
        return

    with open(output_file, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["檔名", "副檔名", "完整路徑", "大小(KB)", "偵測到的個資類型", "處理結果"])

        for item in results:
            writer.writerow([
                item["name"],
                item["ext"],
                item["path"],
                item["size_kb"],
                "、".join(item["pii_types"]),
                item.get("action_result", "")
            ])

    print(f"\n📄 已輸出報表：{output_file}")


def scan_nas() -> None:
    results = run_scan({"nas_smb_root": NAS_PATH}, print)

    print("\n========== 掃描結果 ==========")
    if results:
        for i, item in enumerate(results, start=1):
            print(f"{i}. {item['path']} | 類型: {', '.join(item['pii_types'])}")
        print(f"\n總共找到 {len(results)} 個含個資檔案")
    else:
        print("沒有偵測到含個資的檔案")

    export_csv(results)


if __name__ == "__main__":
    scan_nas()
